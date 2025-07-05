#!/usr/bin/env python3

import json
import os
import glob
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Define the report directory
report_dir = "/tmp/compliance_report"

# Gather all .json files from report directory
data_files = glob.glob(os.path.join(report_dir, "*.json"))

if not data_files:
    print("No JSON data files found in", report_dir)
    exit(1)

# Load data from all files
rows = []
for filepath in data_files:
    with open(filepath) as f:
        try:
            rows.append(json.load(f))
        except Exception as e:
            print(f"Error reading {filepath}: {e}")

# Convert to DataFrame
df = pd.DataFrame(rows)

# Reorder columns (if available)
columns = ["ip", "name", "os", "kernel", "uptime", "compliance"]
df = df[[col for col in columns if col in df.columns]]

# Output Excel file path
excel_file = os.path.join(report_dir, "compliance_report.xlsx")

# Write to Excel with styling
with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Compliance Report", index=False)
    worksheet = writer.sheets["Compliance Report"]

    # Format column widths
    for i, column in enumerate(df.columns, 1):
        max_length = max(df[column].astype(str).map(len).max(), len(column))
        worksheet.column_dimensions[get_column_letter(i)].width = max_length + 4

    # Style header
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

print("✅ Excel report generated:", excel_file)
